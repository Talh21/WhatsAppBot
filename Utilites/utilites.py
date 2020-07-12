import time
import openpyxl as excel
import random
from Utilites.validation import *
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import InvalidArgumentException
from selenium import webdriver
import requests

CONTACTS_AND_MESSAGES_XLSX = "Contacts and messages.xlsx"


def get_user_contacts_using_prompt():
    contacts_list = []
    print('Enter the contact name that you wish to send the message to, when you finish just type "Done"')
    contact = input(str("\rContact name: "))
    contacts_list.append(contact)

    while contact.lower() != "done":
        contact = input(str("\rContact name: "))
        contacts_list.append(contact)
    contacts_list.pop()
    return contacts_list


def get_contacts_from_file(file):
    contacts_list = []
    try:
        file = excel.load_workbook(file)
        sheet = file.active
        first_column = sheet['A']
        for cell in range(1, sheet.max_row):
            if first_column[cell].value is not None:  # checking if cell is not empty in order to append it to the list
                contacts_list.append(first_column[cell].value)

        if len(contacts_list) == 0:  # if all column is empty...
            print("\n[-] Pay attention! Contacts list column in file is empty...\n"
                  "Please enter it manually. ")
            return False
        return contacts_list

    except FileNotFoundError:
        is_user_wants_to_quit = input('File was`nt found! Press Enter to enter contacts manually or press "Q" to exit: ')
        while is_user_wants_to_quit != "" and is_user_wants_to_quit.lower() != "q":
            is_user_wants_to_quit = input('[-] Not a valid option. Try again:  ')

        if is_user_wants_to_quit.lower() == "q":
            print("\nThank you! Hope to see you soon (:")
            quit()
        else:
            contacts_list = get_user_contacts_using_prompt()
            return contacts_list


def get_message():
    take_msg_from_file = input("Do you want to load a random message from a file? [YES/NO]\nEnter your answer: ")
    while check_if_yes_or_no(take_msg_from_file) is False:
        take_msg_from_file = input("You can chose only: [YES/NO]\nTry again:")
    if take_msg_from_file.lower() == "yes":
        rand_msg = get_messages_from_file(CONTACTS_AND_MESSAGES_XLSX)
        if rand_msg is not False:
            return rand_msg
        print("[-] Pay attention! The messages column/file is empty!\nTaking from you a message...")
    else:
        time.sleep(0.5)
        main_message = input("\n\rEnter the main message: ")
        return main_message


def get_messages_from_file(file):
    message_list = []
    try:
        file = excel.load_workbook(file)
        sheet = file.active
        message_column = sheet['B']
        for cell in range(1, sheet.max_row):
            if message_column[cell].value is not None:  # checking if cell is not empty in order to append it to the list
                message_list.append(message_column[cell].value)
        random_message = (random.choice(message_list))
        if len(message_list) == 0:  # if all column is empty...
            return False
        return random_message

    except FileNotFoundError:
        is_user_wants_to_quit = input('File was`nt found! Press Enter to write the main message'
                                      'manually or press "Q" to exit: ')
        while is_user_wants_to_quit.lower() != "" and is_user_wants_to_quit.lower() != "q":
            is_user_wants_to_quit = input('[-] Not a valid option. Try again:  ')

        if is_user_wants_to_quit.lower() == "q":
            print("\nThank you! Hope to see you soon (:")
            quit()
        else:
            main_message = input("\n\rEnter the main message: ")
            return main_message
